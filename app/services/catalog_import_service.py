"""Импорт каталога из выгрузок маркетплейсов (Ozon / Wildberries / Яндекс.Маркет).

Парсит .xlsx-файлы, сопоставляет колонки по alias + keyword-матчингу
(шаблоны выгрузок периодически меняются, поэтому matcher резистентный),
и создаёт товары + варианты в БД.
"""
from io import BytesIO
from typing import Any, Literal

from openpyxl import load_workbook
from sqlalchemy import select

from app.database.db import async_session
from app.models.category import Category
from app.models.product import Product
from app.models.product_variant import ProductVariant

MarketplaceSource = Literal["ozon", "wb", "ym"]

DEFAULT_CATEGORY_NAME = "Импортировано"

# Точные alias'ы названий колонок для каждого источника.
# Список неполный — дополняется keyword-матчингом ниже.
_COLUMN_ALIASES: dict[str, dict[str, list[str]]] = {
    "ozon": {
        "name": ["Название товара", "Название", "Product name", "Наименование"],
        "price": ["Цена, руб.", "Цена", "Цена с учетом скидок и промокодов, руб.", "Цена до скидок, руб."],
        "stock": ["Остаток на складе", "Доступно к продаже", "Остатки", "Свободные остатки"],
        "sku": ["Артикул", "Артикул Ozon", "SKU"],
    },
    "wb": {
        "name": ["Наименование", "Название", "Коммерческое наименование"],
        "price": ["Цена продавца", "Розничная цена", "Цена", "price"],
        "stock": ["Количество", "Остаток", "Кол-во", "К свободной продаже"],
        "sku": ["Артикул продавца", "Артикул цвета", "Артикул", "Баркод"],
    },
    "ym": {
        "name": ["Название", "Наименование товара", "name", "Название товара"],
        "price": ["Цена", "price", "Текущая цена"],
        "stock": ["Остатки", "Количество на складе", "stock", "Кол-во"],
        "sku": ["Артикул", "SKU", "shopSku"],
    },
}

# Keyword-матчинг: если точный alias не найден, ищем подстроку (case-insensitive).
# Порядок важен — "артикул" проверяется до "назв" и т.д.
_COLUMN_KEYWORDS: dict[str, list[str]] = {
    "name": ["назван", "наименован", "product name"],
    "price": ["цен", "price"],
    "stock": ["остат", "количеств", "stock", "кол-во", "склад"],
    "sku": ["артикул", "sku", "баркод"],
}

_ALL_FIELDS = list(_COLUMN_ALIASES["ozon"].keys())

# Сколько первых строк проверять в поисках строки заголовков.
# Выгрузки WB имеют двухуровневый заголовок (группы колонок → имена полей),
# поэтому простого чтения первой строки недостаточно.
HEADER_SCAN_ROWS = 5


def _match_column(header: str, source: MarketplaceSource) -> str | None:
    """Сопоставляет заголовок колонки с полем (name/price/stock/sku).

    Сначала точное совпадение по alias, потом keyword-подстрока.
    Возвращает имя поля или None.
    """
    header_norm = header.strip().lower()

    for field in _ALL_FIELDS:
        for alias in _COLUMN_ALIASES[source].get(field, []):
            if header.strip() == alias:
                return field

    for field, keywords in _COLUMN_KEYWORDS.items():
        for kw in keywords:
            if kw in header_norm:
                return field

    return None


def _parse_price(value: Any) -> int | None:
    """Парсит цену из ячейки — может быть int, float или строкой '1500.00'."""
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return int(round(float(value)))
    s = str(value).strip().replace(" ", "").replace("\xa0", "").replace("₽", "")
    s = s.replace("руб.", "").replace(",", ".")
    try:
        return int(round(float(s)))
    except (ValueError, TypeError):
        return None


def _parse_stock(value: Any) -> int | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return int(value)
    s = str(value).strip().replace(" ", "").replace("\xa0", "")
    try:
        return int(float(s))
    except (ValueError, TypeError):
        return None


class CatalogImportService:
    """Парсинг и импорт каталога из выгрузок маркетплейсов."""

    @staticmethod
    def parse_marketplace_file(file_bytes: bytes, source: MarketplaceSource) -> dict:
        """Парсит .xlsx файл и возвращает превью без записи в БД.

        Возвращает dict:
            source, total_rows, recognized_rows,
            rows: [{row_number, name, price, stock, category_guess, warnings, recognized}],
            unmapped_columns: [str]
        """
        wb = load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
        ws = wb.active

        # Выгрузки маркетплейсов (особенно Wildberries) могут иметь
        # многоуровневые заголовки: первая строка — названия групп колонок
        # («Основная информация», «Габариты» …), вторая — реальные имена полей.
        # Сканируем первые несколько строк и выбираем ту, где больше всего
        # сопоставленных полей — это и есть строка заголовков.
        candidate_rows = list(
            ws.iter_rows(min_row=1, max_row=HEADER_SCAN_ROWS, values_only=True)
        )
        if not candidate_rows:
            wb.close()
            return {
                "source": source,
                "total_rows": 0,
                "recognized_rows": 0,
                "rows": [],
                "unmapped_columns": [],
            }

        best_header_idx = 0
        best_match_count = -1
        column_map: dict[str, int] = {}
        raw_headers: list[str] = []

        for idx, cand_row in enumerate(candidate_rows):
            cand_headers = [
                str(c).strip() if c is not None else "" for c in cand_row
            ]
            cand_map: dict[str, int] = {}
            for col_idx, header in enumerate(cand_headers):
                if not header:
                    continue
                field = _match_column(header, source)
                if field and field not in cand_map:
                    cand_map[field] = col_idx
            if len(cand_map) > best_match_count:
                best_match_count = len(cand_map)
                best_header_idx = idx
                column_map = cand_map
                raw_headers = cand_headers

        header_row_num = best_header_idx + 1  # 1-based

        unmapped_columns: list[str] = []
        for header in raw_headers:
            if not header:
                continue
            if not _match_column(header, source):
                unmapped_columns.append(header)

        name_idx = column_map.get("name")
        price_idx = column_map.get("price")
        stock_idx = column_map.get("stock")

        rows: list[dict] = []

        for row_num, row in enumerate(
            ws.iter_rows(min_row=header_row_num + 1, values_only=True),
            start=header_row_num + 1,
        ):
            if row is None or all(c is None or str(c).strip() == "" for c in row):
                continue

            name = str(row[name_idx]).strip() if name_idx is not None and name_idx < len(row) and row[name_idx] else ""
            price = _parse_price(row[price_idx]) if price_idx is not None and price_idx < len(row) else None
            stock = _parse_stock(row[stock_idx]) if stock_idx is not None and stock_idx < len(row) else None

            warnings: list[str] = []
            recognized = True

            if not name:
                recognized = False
                warnings.append("Не распознано название товара")
            if price is None:
                warnings.append("Не распознана цена")
                recognized = False
            if stock is None:
                warnings.append("Не распознан остаток")

            rows.append({
                "row_number": row_num,
                "name": name,
                "price": price,
                "stock": stock,
                "category_guess": DEFAULT_CATEGORY_NAME,
                "warnings": warnings,
                "recognized": recognized,
            })

        wb.close()

        recognized_count = sum(1 for r in rows if r["recognized"])

        return {
            "source": source,
            "total_rows": len(rows),
            "recognized_rows": recognized_count,
            "rows": rows,
            "unmapped_columns": unmapped_columns,
        }

    @staticmethod
    async def import_rows(
        shop_id: int,
        rows: list[dict],
        category_id: int | None = None,
    ) -> dict:
        """Создаёт Product + ProductVariant для каждой подтверждённой строки.

        Если category_id не указан — создаёт (или находит) категорию
        с именем «Импортировано».

        Возвращает {created, category_id}.
        """
        async with async_session() as session:
            if category_id is None:
                result = await session.execute(
                    select(Category).where(
                        Category.shop_id == shop_id,
                        Category.name == DEFAULT_CATEGORY_NAME,
                    )
                )
                category = result.scalar_one_or_none()
                if category is None:
                    category = Category(shop_id=shop_id, name=DEFAULT_CATEGORY_NAME)
                    session.add(category)
                    await session.flush()
                category_id = category.id

            created = 0
            for row in rows:
                product = Product(
                    shop_id=shop_id,
                    category_id=category_id,
                    name=row["name"],
                    description="",
                    is_active=True,
                )
                product.variants = [
                    ProductVariant(
                        shop_id=shop_id,
                        volume="Стандарт",
                        price=row.get("price") or 0,
                        stock=row.get("stock") or 0,
                        attributes=row.get("attributes") or {},
                    )
                ]
                session.add(product)
                created += 1

            await session.commit()

            return {"created": created, "category_id": category_id}
