from sqlalchemy import select, func
from sqlalchemy.orm import selectinload

from app.database.db import async_session
from app.models.category import Category
from app.models.product import Product
from app.models.product_variant import ProductVariant
from app.models.product_photo import ProductPhoto

from io import BytesIO

from openpyxl import Workbook, load_workbook


def _category_to_dict(category: Category) -> dict:
    return {
        "id": category.id,
        "name": category.name,
        "emoji": category.emoji,
    }


def _product_to_dict(product: Product) -> dict:
    return {
        "id": product.id,
        "category_id": product.category_id,
        "name": product.name,
        "description": product.description,
        "is_active": product.is_active,
        "variants": [
            {
                "id": variant.id,
                "volume": variant.volume,
                "price": variant.price,
                "stock": variant.stock,
                "attributes": variant.attributes or {},
            }
            for variant in product.variants
        ],
        "photos": [
            {
                "id": photo.id,
                "file_id": photo.file_id,
                "position": photo.position,
            }
            for photo in product.photos
        ],
    }


class CatalogAdminService:
    """
    Управление каталогом: категории, товары, варианты, фото.
    В отличие от CatalogService, видит все товары — включая скрытые.
    """

    # ==========================
    # Категории
    # ==========================

    @staticmethod
    async def get_categories(shop_id: int) -> list[dict]:
        async with async_session() as session:
            result = await session.execute(
                select(Category)
                .where(Category.shop_id == shop_id)
                .order_by(Category.id)
            )
            return [_category_to_dict(c) for c in result.scalars().all()]

    @staticmethod
    async def create_category(shop_id: int, name: str, emoji: str | None = None) -> int:
        async with async_session() as session:
            category = Category(shop_id=shop_id, name=name, emoji=emoji or None)
            session.add(category)
            await session.commit()
            await session.refresh(category)
            return category.id

    @staticmethod
    async def rename_category(shop_id: int, category_id: int, name: str) -> None:
        async with async_session() as session:
            result = await session.execute(
                select(Category).where(
                    Category.shop_id == shop_id,
                    Category.id == category_id,
                )
            )
            category = result.scalar_one_or_none()
            if category:
                category.name = name
                await session.commit()

    @staticmethod
    async def update_category_emoji(shop_id: int, category_id: int, emoji: str | None) -> None:
        async with async_session() as session:
            result = await session.execute(
                select(Category).where(
                    Category.shop_id == shop_id,
                    Category.id == category_id,
                )
            )
            category = result.scalar_one_or_none()
            if category:
                category.emoji = emoji or None
                await session.commit()

    @staticmethod
    async def count_products_in_category(shop_id: int, category_id: int) -> int:
        async with async_session() as session:
            result = await session.execute(
                select(func.count())
                .select_from(Product)
                .where(
                    Product.shop_id == shop_id,
                    Product.category_id == category_id,
                )
            )
            return result.scalar() or 0

    @staticmethod
    async def delete_category(shop_id: int, category_id: int) -> bool:
        """
        Удаляет категорию.
        Возвращает False, если в категории есть товары.
        """
        count = await CatalogAdminService.count_products_in_category(shop_id, category_id)

        if count > 0:
            return False

        async with async_session() as session:
            result = await session.execute(
                select(Category).where(
                    Category.shop_id == shop_id,
                    Category.id == category_id,
                )
            )
            category = result.scalar_one_or_none()

            if category:
                await session.delete(category)
                await session.commit()

        return True

    # ==========================
    # Товары
    # ==========================

    @staticmethod
    async def get_products(shop_id: int, category_id: int) -> list[dict]:
        async with async_session() as session:
            result = await session.execute(
                select(Product)
                .options(
                    selectinload(Product.variants),
                    selectinload(Product.photos),
                )
                .where(
                    Product.shop_id == shop_id,
                    Product.category_id == category_id,
                )
                .order_by(Product.id)
            )
            return [_product_to_dict(p) for p in result.scalars().all()]

    @staticmethod
    async def get_product(shop_id: int, product_id: int) -> dict | None:
        async with async_session() as session:
            result = await session.execute(
                select(Product)
                .options(
                    selectinload(Product.variants),
                    selectinload(Product.photos),
                )
                .where(
                    Product.shop_id == shop_id,
                    Product.id == product_id,
                )
            )
            product = result.scalar_one_or_none()
            return _product_to_dict(product) if product else None

    @staticmethod
    async def create_product(
        shop_id: int,
        category_id: int,
        name: str,
        description: str,
        variants: list[dict],
        photos: list[str] | None = None,
    ) -> int:
        async with async_session() as session:
            product = Product(
                shop_id=shop_id,
                category_id=category_id,
                name=name,
                description=description,
                is_active=True,
            )

            product.variants = [
                ProductVariant(
                    shop_id=shop_id,
                    volume=variant["volume"],
                    price=variant["price"],
                    stock=variant.get("stock", 0),
                    attributes=variant.get("attributes") or {},
                )
                for variant in variants
            ]

            if photos:
                product.photos = [
                    ProductPhoto(shop_id=shop_id, file_id=file_id, position=i)
                    for i, file_id in enumerate(photos)
                ]

            session.add(product)
            await session.commit()
            await session.refresh(product)

            return product.id

    @staticmethod
    async def delete_product(shop_id: int, product_id: int) -> None:
        async with async_session() as session:
            result = await session.execute(
                select(Product).where(
                    Product.shop_id == shop_id,
                    Product.id == product_id,
                )
            )
            product = result.scalar_one_or_none()
            if product:
                await session.delete(product)
                await session.commit()

    @staticmethod
    async def toggle_active(shop_id: int, product_id: int) -> bool | None:
        async with async_session() as session:
            result = await session.execute(
                select(Product).where(
                    Product.shop_id == shop_id,
                    Product.id == product_id,
                )
            )
            product = result.scalar_one_or_none()

            if not product:
                return None

            product.is_active = not product.is_active
            await session.commit()

            return product.is_active

    @staticmethod
    async def set_active(shop_id: int, product_id: int, is_active: bool) -> None:
        async with async_session() as session:
            result = await session.execute(
                select(Product).where(
                    Product.shop_id == shop_id,
                    Product.id == product_id,
                )
            )
            product = result.scalar_one_or_none()

            if not product:
                return

            product.is_active = is_active
            await session.commit()

    @staticmethod
    async def update_product(
        shop_id: int,
        product_id: int,
        name: str | None = None,
        description: str | None = None,
    ) -> None:
        async with async_session() as session:
            result = await session.execute(
                select(Product).where(
                    Product.shop_id == shop_id,
                    Product.id == product_id,
                )
            )
            product = result.scalar_one_or_none()

            if not product:
                return

            if name is not None:
                product.name = name
            if description is not None:
                product.description = description

            await session.commit()

    @staticmethod
    async def update_variant_stock(shop_id: int, variant_id: int, stock: int) -> bool:
        async with async_session() as session:
            result = await session.execute(
                select(ProductVariant).where(
                    ProductVariant.shop_id == shop_id,
                    ProductVariant.id == variant_id,
                )
            )
            variant = result.scalar_one_or_none()

            if not variant:
                return False

            variant.stock = max(0, stock)
            await session.commit()
            return True

    @staticmethod
    async def update_variant(
        shop_id: int,
        variant_id: int,
        volume: str | None = None,
        price: int | None = None,
        stock: int | None = None,
        attributes: dict | None = None,
    ) -> bool:
        async with async_session() as session:
            result = await session.execute(
                select(ProductVariant).where(
                    ProductVariant.shop_id == shop_id,
                    ProductVariant.id == variant_id,
                )
            )
            variant = result.scalar_one_or_none()

            if not variant:
                return False

            if volume is not None:
                variant.volume = volume
            if price is not None:
                variant.price = price
            if stock is not None:
                variant.stock = max(0, stock)
            if attributes is not None:
                variant.attributes = attributes

            await session.commit()
            return True

    @staticmethod
    async def add_variant(
        shop_id: int,
        product_id: int,
        volume: str,
        price: int,
        stock: int = 0,
        attributes: dict | None = None,
    ) -> int | None:
        async with async_session() as session:
            variant = ProductVariant(
                shop_id=shop_id,
                product_id=product_id,
                volume=volume,
                price=price,
                stock=stock,
                attributes=attributes or {},
            )
            session.add(variant)
            await session.commit()
            await session.refresh(variant)
            return variant.id

    @staticmethod
    async def delete_variant(shop_id: int, variant_id: int) -> bool:
        async with async_session() as session:
            result = await session.execute(
                select(ProductVariant).where(
                    ProductVariant.shop_id == shop_id,
                    ProductVariant.id == variant_id,
                )
            )
            variant = result.scalar_one_or_none()
            if not variant:
                return False
            await session.delete(variant)
            await session.commit()
            return True

    @staticmethod
    async def add_photo(shop_id: int, product_id: int, file_id: str) -> int | None:
        async with async_session() as session:
            result = await session.execute(
                select(func.count())
                .select_from(ProductPhoto)
                .where(
                    ProductPhoto.shop_id == shop_id,
                    ProductPhoto.product_id == product_id,
                )
            )
            position = result.scalar() or 0

            photo = ProductPhoto(
                shop_id=shop_id,
                product_id=product_id,
                file_id=file_id,
                position=position,
            )
            session.add(photo)
            await session.commit()
            await session.refresh(photo)

            return photo.id

    @staticmethod
    async def delete_photo(shop_id: int, photo_id: int) -> None:
        async with async_session() as session:
            result = await session.execute(
                select(ProductPhoto).where(
                    ProductPhoto.shop_id == shop_id,
                    ProductPhoto.id == photo_id,
                )
            )
            photo = result.scalar_one_or_none()
            if photo:
                await session.delete(photo)
                await session.commit()

    # ==========================
    # Все товары (вне категорий)
    # ==========================

    @staticmethod
    async def get_all_products(shop_id: int) -> list[dict]:
        async with async_session() as session:
            result = await session.execute(
                select(Product)
                .options(
                    selectinload(Product.variants),
                    selectinload(Product.photos),
                    selectinload(Product.category),
                )
                .where(Product.shop_id == shop_id)
                .order_by(Product.id.desc())
            )
            products = result.scalars().all()

            return [
                {
                    **_product_to_dict(p),
                    "category_name": p.category.name if p.category else None,
                }
                for p in products
            ]

    # ==========================
    # Массовое обновление остатков
    # ==========================

    @staticmethod
    async def get_stock_template_data(shop_id: int) -> list[dict]:
        """Возвращает данные всех вариантов для шаблона остатков."""
        async with async_session() as session:
            result = await session.execute(
                select(Product, ProductVariant)
                .join(ProductVariant, ProductVariant.product_id == Product.id)
                .where(Product.shop_id == shop_id)
                .order_by(Product.id, ProductVariant.id)
            )
            rows = result.all()
            return [
                {
                    "variant_id": v.id,
                    "product_name": p.name,
                    "variant_volume": v.volume,
                    "current_stock": v.stock,
                }
                for p, v in rows
            ]

    @staticmethod
    def generate_stock_template_xlsx(data: list[dict]) -> bytes:
        """Генерирует .xlsx-шаблон для массового обновления остатков.

        Колонки: id (скрытый), Название, Вариант, Текущий остаток, Остаток.
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Остатки"
        ws.append(["id", "Название", "Вариант", "Текущий остаток", "Остаток"])

        ws.column_dimensions["A"].hidden = True

        for row in data:
            ws.append([
                row["variant_id"],
                row["product_name"],
                row["variant_volume"],
                row["current_stock"],
                None,
            ])

        buf = BytesIO()
        wb.save(buf)
        wb.close()
        return buf.getvalue()

    @staticmethod
    def parse_stock_file(file_bytes: bytes) -> dict:
        """Парсит заполненный шаблон остатков.

        Возвращает {updates: [{variant_id, stock}], errors: [str]}.
        """
        wb = load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
        ws = wb.active

        rows = list(ws.iter_rows(min_row=1, values_only=True))
        wb.close()

        if not rows:
            return {"updates": [], "errors": ["Файл пустой"]}

        header = [str(c).strip().lower() if c is not None else "" for c in rows[0]]

        id_idx = None
        stock_idx = None
        for i, h in enumerate(header):
            if h == "id":
                id_idx = i
            elif h in ("остаток", "новый остаток"):
                stock_idx = i

        if id_idx is None:
            return {"updates": [], "errors": ["Колонка «id» не найдена. Используйте шаблон, скачанный из админ-панели."]}
        if stock_idx is None:
            return {"updates": [], "errors": ["Колонка «Остаток» не найдена. Используйте шаблон, скачанный из админ-панели."]}

        updates: list[dict] = []
        errors: list[str] = []

        for row_num, row in enumerate(rows[1:], start=2):
            if row is None or all(c is None or str(c).strip() == "" for c in row):
                continue

            raw_id = row[id_idx] if id_idx < len(row) else None
            raw_stock = row[stock_idx] if stock_idx < len(row) else None

            if raw_id is None or str(raw_id).strip() == "":
                continue

            try:
                variant_id = int(float(str(raw_id).strip()))
            except (ValueError, TypeError):
                errors.append(f"Строка {row_num}: id «{raw_id}» не является числом")
                continue

            if raw_stock is None or str(raw_stock).strip() == "":
                continue

            try:
                stock = int(float(str(raw_stock).strip()))
            except (ValueError, TypeError):
                errors.append(f"Строка {row_num}: остаток «{raw_stock}» не является числом")
                continue

            updates.append({"variant_id": variant_id, "stock": max(0, stock)})

        return {"updates": updates, "errors": errors}

    @staticmethod
    async def apply_stock_updates(shop_id: int, updates: list[dict]) -> dict:
        """Применяет массовое обновление остатков одним batch-запросом.

        Возвращает {updated, not_found}.
        """
        if not updates:
            return {"updated": 0, "not_found": 0}

        variant_ids = [u["variant_id"] for u in updates]
        stock_map = {u["variant_id"]: u["stock"] for u in updates}

        async with async_session() as session:
            result = await session.execute(
                select(ProductVariant).where(
                    ProductVariant.shop_id == shop_id,
                    ProductVariant.id.in_(variant_ids),
                )
            )
            variants = result.scalars().all()

            found_ids = set()
            for v in variants:
                v.stock = stock_map[v.id]
                found_ids.add(v.id)

            await session.commit()

        not_found = len(variant_ids) - len(found_ids)
        return {"updated": len(found_ids), "not_found": not_found}
