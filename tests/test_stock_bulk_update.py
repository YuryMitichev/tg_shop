"""Тесты массового обновления остатков через .xlsx-шаблон."""
from datetime import datetime, timedelta, timezone
from io import BytesIO
from unittest.mock import AsyncMock, patch

import pytest
from httpx import ASGITransport, AsyncClient
from openpyxl import Workbook, load_workbook

from app.api.main import create_app
from app.models.subscription import Subscription, SubscriptionPlan
from app.services.admin_auth_service import AdminAuthService
from app.services.catalog_admin_service import CatalogAdminService
from app.utils.crypto import encrypt


_ADMIN_DICT = {"admin_id": 123456, "shop_id": 1, "is_super_admin": False}


@pytest.fixture
def admin_cookie():
    token = AdminAuthService._create_token(123456, shop_id=1, is_super_admin=False)
    return {"admin_token": token}


@pytest.fixture
def mock_admin_auth():
    with patch(
        "app.api.admin_auth.AdminAuthService.verify_token",
        new_callable=AsyncMock,
        return_value=_ADMIN_DICT,
    ):
        yield


@pytest.fixture
async def active_subscription(db_session):
    session_maker = db_session
    now = datetime.now(timezone.utc).replace(tzinfo=None)
    async with session_maker() as session:
        session.add(SubscriptionPlan(
            id=1, name="Тест", price=5000, duration_days=30, is_trial=False,
        ))
        await session.commit()
        session.add(Subscription(
            shop_id=1, plan_id=1, status="active",
            started_at=now, expires_at=now + timedelta(days=25),
        ))
        await session.commit()


# ==========================
# CatalogAdminService: get_stock_template_data
# ==========================

class TestGetStockTemplateData:

    async def test_returns_all_variants(self, db_session, seed_data):
        data = await CatalogAdminService.get_stock_template_data(1)

        assert len(data) == 4

        v1 = data[0]
        assert v1["variant_id"] == 1
        assert v1["product_name"] == "Кашемир"
        assert v1["variant_volume"] == "75 г"
        assert v1["current_stock"] == 10

    async def test_empty_shop(self, db_session, seed_data):
        data = await CatalogAdminService.get_stock_template_data(999)
        assert data == []


# ==========================
# CatalogAdminService: generate_stock_template_xlsx
# ==========================

class TestGenerateStockTemplateXlsx:

    def test_generates_valid_xlsx_with_headers(self):
        data = [
            {"variant_id": 1, "product_name": "Товар A", "variant_volume": "100г", "current_stock": 5},
            {"variant_id": 2, "product_name": "Товар B", "variant_volume": "200г", "current_stock": 0},
        ]

        xlsx_bytes = CatalogAdminService.generate_stock_template_xlsx(data)

        wb = load_workbook(BytesIO(xlsx_bytes))
        ws = wb.active

        headers = [c.value for c in ws[1]]
        assert headers == ["id", "Название", "Вариант", "Текущий остаток", "Остаток"]

        assert ws.cell(row=2, column=1).value == 1
        assert ws.cell(row=2, column=2).value == "Товар A"
        assert ws.cell(row=2, column=3).value == "100г"
        assert ws.cell(row=2, column=4).value == 5
        assert ws.cell(row=2, column=5).value is None

        assert ws.cell(row=3, column=1).value == 2
        wb.close()

    def test_column_a_is_hidden(self):
        data = [{"variant_id": 1, "product_name": "X", "variant_volume": "Y", "current_stock": 0}]

        xlsx_bytes = CatalogAdminService.generate_stock_template_xlsx(data)

        wb = load_workbook(BytesIO(xlsx_bytes))
        ws = wb.active
        assert ws.column_dimensions["A"].hidden is True
        wb.close()

    def test_empty_data(self):
        xlsx_bytes = CatalogAdminService.generate_stock_template_xlsx([])

        wb = load_workbook(BytesIO(xlsx_bytes))
        ws = wb.active
        headers = [c.value for c in ws[1]]
        assert headers == ["id", "Название", "Вариант", "Текущий остаток", "Остаток"]
        assert ws.max_row == 1
        wb.close()


# ==========================
# CatalogAdminService: parse_stock_file
# ==========================

class TestParseStockFile:

    def _make_xlsx(self, rows: list[list]) -> bytes:
        wb = Workbook()
        ws = wb.active
        ws.append(["id", "Название", "Вариант", "Текущий остаток", "Остаток"])
        for r in rows:
            ws.append(r)
        buf = BytesIO()
        wb.save(buf)
        wb.close()
        return buf.getvalue()

    def test_parses_valid_file(self):
        xlsx = self._make_xlsx([
            [1, "Товар A", "100г", 5, 10],
            [2, "Товар B", "200г", 0, 3],
        ])

        result = CatalogAdminService.parse_stock_file(xlsx)

        assert result["errors"] == []
        assert len(result["updates"]) == 2
        assert result["updates"][0] == {"variant_id": 1, "stock": 10}
        assert result["updates"][1] == {"variant_id": 2, "stock": 3}

    def test_skips_empty_stock_rows(self):
        xlsx = self._make_xlsx([
            [1, "Товар A", "100г", 5, 10],
            [2, "Товар B", "200г", 0, None],
        ])

        result = CatalogAdminService.parse_stock_file(xlsx)

        assert len(result["updates"]) == 1
        assert result["updates"][0]["variant_id"] == 1

    def test_skips_empty_rows(self):
        xlsx = self._make_xlsx([
            [1, "Товар A", "100г", 5, 10],
            [None, None, None, None, None],
        ])

        result = CatalogAdminService.parse_stock_file(xlsx)

        assert len(result["updates"]) == 1

    def test_negative_stock_clamped_to_zero(self):
        xlsx = self._make_xlsx([
            [1, "Товар A", "100г", 5, -3],
        ])

        result = CatalogAdminService.parse_stock_file(xlsx)

        assert result["updates"][0]["stock"] == 0

    def test_missing_id_column(self):
        wb = Workbook()
        ws = wb.active
        ws.append(["Название", "Остаток"])
        ws.append(["Товар A", 10])
        buf = BytesIO()
        wb.save(buf)
        wb.close()

        result = CatalogAdminService.parse_stock_file(buf.getvalue())

        assert len(result["updates"]) == 0
        assert any("id" in e for e in result["errors"])

    def test_missing_stock_column(self):
        wb = Workbook()
        ws = wb.active
        ws.append(["id", "Название"])
        ws.append([1, "Товар A"])
        buf = BytesIO()
        wb.save(buf)
        wb.close()

        result = CatalogAdminService.parse_stock_file(buf.getvalue())

        assert len(result["updates"]) == 0
        assert any("Остаток" in e for e in result["errors"])

    def test_invalid_stock_value_error(self):
        xlsx = self._make_xlsx([
            [1, "Товар A", "100г", 5, "abc"],
        ])

        result = CatalogAdminService.parse_stock_file(xlsx)

        assert len(result["updates"]) == 0
        assert any("не является числом" in e for e in result["errors"])

    def test_float_stock_truncated(self):
        xlsx = self._make_xlsx([
            [1, "Товар A", "100г", 5, 7.9],
        ])

        result = CatalogAdminService.parse_stock_file(xlsx)

        assert result["updates"][0]["stock"] == 7


# ==========================
# CatalogAdminService: apply_stock_updates
# ==========================

class TestApplyStockUpdates:

    async def test_updates_stock(self, db_session, seed_data):
        from sqlalchemy import select
        from app.models.product_variant import ProductVariant

        updates = [
            {"variant_id": 1, "stock": 100},
            {"variant_id": 2, "stock": 50},
        ]

        result = await CatalogAdminService.apply_stock_updates(1, updates)

        assert result["updated"] == 2
        assert result["not_found"] == 0

        async with db_session() as session:
            v1 = (await session.execute(
                select(ProductVariant).where(ProductVariant.id == 1)
            )).scalar_one()
            assert v1.stock == 100

    async def test_not_found_variants(self, db_session, seed_data):
        updates = [
            {"variant_id": 1, "stock": 10},
            {"variant_id": 999, "stock": 20},
        ]

        result = await CatalogAdminService.apply_stock_updates(1, updates)

        assert result["updated"] == 1
        assert result["not_found"] == 1

    async def test_cross_shop_isolated(self, db_session, seed_data):
        from app.services.shop_service import ShopService

        await ShopService.create("Магазин 2", "tok:xshop", 222)

        updates = [
            {"variant_id": 1, "stock": 999},
        ]

        result = await CatalogAdminService.apply_stock_updates(2, updates)

        assert result["updated"] == 0
        assert result["not_found"] == 1

    async def test_empty_updates(self, db_session, seed_data):
        result = await CatalogAdminService.apply_stock_updates(1, [])
        assert result == {"updated": 0, "not_found": 0}


# ==========================
# Admin API: GET /catalog/stock-template
# ==========================

class TestStockTemplateEndpoint:

    async def test_download_template(
        self, db_session, seed_data, admin_cookie, mock_admin_auth, active_subscription
    ):
        app = create_app()
        transport = ASGITransport(app=app)
        async with AsyncClient(transport=transport, base_url="http://test") as client:
            resp = await client.get(
                "/api/admin/catalog/stock-template",
                cookies=admin_cookie,
            )

        assert resp.status_code == 200
        assert "spreadsheet" in resp.headers.get("content-type", "")
        assert "attachment" in resp.headers.get("content-disposition", "")

        wb = load_workbook(BytesIO(resp.content))
        ws = wb.active
        headers = [c.value for c in ws[1]]
        assert headers == ["id", "Название", "Вариант", "Текущий остаток", "Остаток"]
        wb.close()

    async def test_requires_auth(self, db_session, seed_data):
        app = create_app()
        transport = ASGITransport(app=app)
        async with AsyncClient(transport=transport, base_url="http://test") as client:
            resp = await client.get("/api/admin/catalog/stock-template")

        assert resp.status_code == 401


# ==========================
# Admin API: POST /catalog/stock/bulk-update
# ==========================

class TestBulkUpdateEndpoint:

    async def test_bulk_update(
        self, db_session, seed_data, admin_cookie, mock_admin_auth, active_subscription
    ):
        from sqlalchemy import select
        from app.models.product_variant import ProductVariant

        wb = Workbook()
        ws = wb.active
        ws.append(["id", "Название", "Вариант", "Текущий остаток", "Остаток"])
        ws.append([1, "Кашемир", "75 г", 10, 100])
        ws.append([2, "Кашемир", "200 г", 5, 50])
        buf = BytesIO()
        wb.save(buf)
        wb.close()

        app = create_app()
        transport = ASGITransport(app=app)
        async with AsyncClient(transport=transport, base_url="http://test") as client:
            resp = await client.post(
                "/api/admin/catalog/stock/bulk-update",
                cookies=admin_cookie,
                files={"file": ("stock.xlsx", buf.getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            )

        assert resp.status_code == 200
        data = resp.json()
        assert data["updated"] == 2
        assert data["not_found"] == 0

        async with db_session() as session:
            v1 = (await session.execute(
                select(ProductVariant).where(ProductVariant.id == 1)
            )).scalar_one()
            assert v1.stock == 100

    async def test_bulk_update_partial_not_found(
        self, db_session, seed_data, admin_cookie, mock_admin_auth, active_subscription
    ):
        wb = Workbook()
        ws = wb.active
        ws.append(["id", "Название", "Вариант", "Текущий остаток", "Остаток"])
        ws.append([1, "A", "v", 10, 5])
        ws.append([9999, "X", "v", 0, 1])
        buf = BytesIO()
        wb.save(buf)
        wb.close()

        app = create_app()
        transport = ASGITransport(app=app)
        async with AsyncClient(transport=transport, base_url="http://test") as client:
            resp = await client.post(
                "/api/admin/catalog/stock/bulk-update",
                cookies=admin_cookie,
                files={"file": ("stock.xlsx", buf.getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            )

        assert resp.status_code == 200
        data = resp.json()
        assert data["updated"] == 1
        assert data["not_found"] == 1

    async def test_requires_auth(self, db_session, seed_data):
        app = create_app()
        transport = ASGITransport(app=app)
        async with AsyncClient(transport=transport, base_url="http://test") as client:
            resp = await client.post("/api/admin/catalog/stock/bulk-update")

        assert resp.status_code == 401


# ==========================
# Import confirm returns stock_template_url
# ==========================

class TestImportConfirmReturnsTemplateUrl:

    async def test_confirm_returns_stock_template_url(
        self, db_session, seed_data, admin_cookie, mock_admin_auth, active_subscription
    ):
        app = create_app()
        transport = ASGITransport(app=app)
        async with AsyncClient(transport=transport, base_url="http://test") as client:
            resp = await client.post(
                "/api/admin/catalog/import/confirm",
                cookies=admin_cookie,
                json={
                    "rows": [
                        {"name": "Товар 1", "description": "", "category": ""},
                    ],
                },
            )

        assert resp.status_code == 200
        data = resp.json()
        assert data["created"] == 1
        assert data["stock_template_url"] == "/catalog/stock-template"
